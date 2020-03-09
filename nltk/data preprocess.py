import nltk
import csv
from openpyxl import load_workbook
from nltk.tokenize import punkt
from nltk.tokenize import sent_tokenize,word_tokenize
from nltk.corpus import stopwords
from nltk.stem import PorterStemmer
class preprocess:
    def func(self):
        ps=PorterStemmer()
        path="D:\Rajesh\googleplaystore_user_reviews.xlsx"
        workbook_obj=load_workbook(path)
        sheet_obj=workbook_obj.active
        max_row=sheet_obj.max_row
        review={}
        rev=[]
        rev1=[]

        for i in range(2,max_row+1):
            x=sheet_obj.cell(row=i, column=1)
            rev.append(x.value)

        for i in range(2,max_row + 1):
            y = sheet_obj.cell(row=i, column=2)
            rev1.append(y.value)

        review=dict(zip(rev,rev1))
        mod_review=[]
        for data in review.values():
            words=word_tokenize(data)
            #phrases=sent_tokenize(data)
            stop=set(stopwords.words('english'))
            #print(stop)
            #print(len(stop))
            filter=[]
            for w in words:
                if w not in stop:
                    filter.append(w)
            stem=[]
            for s in filter:
                 stem=ps.stem(s)
            mod_review.append(stem)
        for dta in mod_review.:
             print(dta)


        document = 'Today the Netherlands celebrates King\'s Day. To honor this tradition, the Dutch embassy in San Francisco invited me to'
        sentences = nltk.sent_tokenize(document)
        data=[]
        for sent in sentences:
            data=data+nltk.pos_tag(nltk.word_tokenize(sent))
        print(data)
        for w in data:
            if 'DT' in w[1]:
                print(w)
ob=preprocess()
ob.func()